import base64
import hashlib
import hmac
import json
import os
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook


HEADERS = [
    "user_id",
    "name",
    "email",
    "password_hash",
    "face_embedding",
    "face_model",
    "face_enrolled_at",
    "created_at",
]


def encode_bytes(value):
    if value is None:
        return ""
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def decode_bytes(value):
    if not value:
        return b""
    padded = value + "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode(padded.encode("ascii"))


def hash_password(password):
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 240000)
    return f"pbkdf2_sha256${encode_bytes(salt)}${encode_bytes(digest)}"


def verify_password(password, stored_hash):
    try:
        algorithm, salt_text, digest_text = stored_hash.split("$", 2)
    except ValueError:
        return False

    if algorithm != "pbkdf2_sha256":
        return False

    salt = decode_bytes(salt_text)
    expected = decode_bytes(digest_text)
    actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 240000)
    return hmac.compare_digest(actual, expected)


class ExcelUserStore:
    def __init__(self, path):
        self.path = path

    @staticmethod
    def decode_bytes(value):
        return decode_bytes(value)

    def ensure_workbook(self):
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        if os.path.exists(self.path):
            workbook = load_workbook(self.path)
            sheet = workbook["users"]
            headers = self._headers(sheet)
            changed = False
            for header in HEADERS:
                if header not in headers:
                    sheet.cell(row=1, column=sheet.max_column + 1, value=header)
                    changed = True
            if changed:
                workbook.save(self.path)
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "users"
        sheet.append(HEADERS)
        workbook.save(self.path)

    def create_user(self, name, email, password_hash):
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook["users"]
        headers = self._headers(sheet)
        user = {
            "user_id": str(uuid.uuid4()),
            "name": name,
            "email": email.strip().lower(),
            "password_hash": password_hash,
            "face_embedding": "",
            "face_model": "",
            "face_enrolled_at": "",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        sheet.append([user.get(header, "") for header in headers])
        workbook.save(self.path)
        return user

    def get_user_by_email(self, email):
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook["users"]
        headers = self._headers(sheet)
        email = email.strip().lower()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            user = dict(zip(headers, row))
            if (user.get("email") or "").strip().lower() == email:
                return {key: ("" if value is None else value) for key, value in user.items()}

        return None

    def save_face_embedding(self, email, embedding, model_name):
        self._update_user(email, {
            "face_embedding": json.dumps(embedding),
            "face_model": model_name,
            "face_enrolled_at": datetime.now(timezone.utc).isoformat(),
        })

    def load_face_embedding(self, email):
        user = self.get_user_by_email(email)
        if not user or not user.get("face_embedding"):
            return None
        return json.loads(user["face_embedding"])

    def users_with_face_embeddings(self):
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook["users"]
        headers = self._headers(sheet)
        users = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            user = {key: ("" if value is None else value) for key, value in dict(zip(headers, row)).items()}
            if user.get("face_embedding"):
                user["face_embedding"] = json.loads(user["face_embedding"])
                users.append(user)

        return users

    def _update_user(self, email, updates):
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook["users"]
        indexes = {header: index + 1 for index, header in enumerate(self._headers(sheet))}
        email = email.strip().lower()

        for row_index in range(2, sheet.max_row + 1):
            current_email = (sheet.cell(row=row_index, column=indexes["email"]).value or "").strip().lower()
            if current_email != email:
                continue

            for key, value in updates.items():
                sheet.cell(row=row_index, column=indexes[key], value=value)
            workbook.save(self.path)
            return

        raise ValueError(f"User not found: {email}")

    @staticmethod
    def _headers(sheet):
        return [cell.value for cell in sheet[1] if cell.value]
