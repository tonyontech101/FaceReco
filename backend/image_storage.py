"""
Image Database Storage Layer
Manages image metadata and feature embeddings in Excel workbook.
"""

import json
import os
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Optional

import numpy as np
from openpyxl import Workbook, load_workbook
from sklearn.metrics.pairwise import cosine_similarity


# Database schema
HEADERS = [
    "image_id",
    "filename",
    "object_name",
    "category",
    "tags",
    "file_path",
    "embedding",
    "created_at",
]


class ExcelImageStore:
    """Excel-based storage for image metadata and embeddings."""
    
    def __init__(self, path: str):
        """
        Initialize image store.
        
        Args:
            path: Path to Excel workbook file
        """
        self.path = path
    
    def ensure_workbook(self):
        """Create workbook with schema if it doesn't exist."""
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        
        if os.path.exists(self.path):
            # Workbook exists - verify it has correct headers
            try:
                workbook = load_workbook(self.path)
            except PermissionError:
                raise PermissionError(
                    f"Cannot access {self.path} - file is open in another program. "
                    f"Please close Excel or any program viewing this file."
                )
            except Exception as e:
                raise Exception(f"Cannot open {self.path}: {e}")
            
            # Check if 'images' sheet exists
            if 'images' not in workbook.sheetnames:
                # Create sheet if missing
                sheet = workbook.create_sheet('images')
                sheet.append(HEADERS)
                try:
                    workbook.save(self.path)
                except PermissionError:
                    raise PermissionError(
                        f"Cannot save {self.path} - file may be open in another program."
                    )
                return
            
            sheet = workbook['images']
            existing_headers = [cell.value for cell in sheet[1] if cell.value]
            
            # Add missing headers
            changed = False
            for header in HEADERS:
                if header not in existing_headers:
                    sheet.cell(row=1, column=sheet.max_column + 1, value=header)
                    changed = True
            
            if changed:
                try:
                    workbook.save(self.path)
                except PermissionError:
                    raise PermissionError(
                        f"Cannot save {self.path} - file may be open in another program."
                    )
            return
        
        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'images'
        sheet.append(HEADERS)
        try:
            workbook.save(self.path)
        except PermissionError:
            raise PermissionError(
                f"Cannot create {self.path} - check file permissions or close any program viewing this file."
            )
    
    def add_image(
        self,
        filename: str,
        object_name: str,
        category: str,
        tags: List[str],
        file_path: str,
        embedding: np.ndarray
    ) -> Dict:
        """
        Add new image to database.
        
        Args:
            filename: Original filename
            object_name: Human-readable object name (e.g., "Dog")
            category: Broad category (e.g., "Animal")
            tags: List of tags for categorization
            file_path: Relative path to image file
            embedding: 1280-dimensional feature vector
            
        Returns:
            Dict with image metadata including generated image_id
        """
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook['images']
        headers = self._headers(sheet)
        
        # Create image record
        image_id = str(uuid.uuid4())
        image_data = {
            "image_id": image_id,
            "filename": filename,
            "object_name": object_name,
            "category": category,
            "tags": json.dumps(tags),
            "file_path": file_path,
            "embedding": json.dumps(embedding.tolist()),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        
        # Append row
        sheet.append([image_data.get(header, "") for header in headers])
        workbook.save(self.path)
        
        return image_data
    
    def get_all_images(self) -> List[Dict]:
        """
        Retrieve all images from database.
        
        Returns:
            List of image records with embeddings as numpy arrays
        """
        self.ensure_workbook()
        workbook = load_workbook(self.path)
        sheet = workbook['images']
        headers = self._headers(sheet)
        
        images = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            image = dict(zip(headers, row))
            
            # Handle None values
            image = {key: ("" if value is None else value) for key, value in image.items()}
            
            # Skip empty rows
            if not image.get("image_id"):
                continue
            
            # Parse JSON fields
            if image.get("tags"):
                try:
                    image["tags"] = json.loads(image["tags"])
                except json.JSONDecodeError:
                    image["tags"] = []
            else:
                image["tags"] = []
            
            if image.get("embedding"):
                try:
                    image["embedding"] = np.array(json.loads(image["embedding"]), dtype=np.float32)
                except (json.JSONDecodeError, ValueError):
                    # Skip images with invalid embeddings
                    continue
            else:
                continue
            
            images.append(image)
        
        return images
    
    def get_image_by_id(self, image_id: str) -> Optional[Dict]:
        """
        Retrieve single image by ID.
        
        Args:
            image_id: Image UUID
            
        Returns:
            Image record or None if not found
        """
        images = self.get_all_images()
        for image in images:
            if image.get("image_id") == image_id:
                return image
        return None
    
    def search_similar(
        self,
        query_embedding: np.ndarray,
        threshold: float = 0.5,
        limit: int = 3
    ) -> List[Dict]:
        """
        Search for similar images using cosine similarity.
        
        Args:
            query_embedding: Feature vector to search for (1280-dim)
            threshold: Minimum similarity score (0-1)
            limit: Maximum number of results to return
            
        Returns:
            List of image records with similarity scores, sorted by similarity (highest first)
            Each record includes a "similarity" field
        """
        images = self.get_all_images()
        
        if not images:
            return []
        
        # Extract all embeddings into a matrix
        embeddings_matrix = np.array([img["embedding"] for img in images])
        
        # Reshape query embedding to 2D array for sklearn
        query_embedding_2d = query_embedding.reshape(1, -1)
        
        # Compute cosine similarity for all images at once (vectorized operation)
        similarities = cosine_similarity(query_embedding_2d, embeddings_matrix)[0]
        
        # Add similarity scores to image records
        for i, image in enumerate(images):
            image["similarity"] = float(similarities[i])
        
        # Filter by threshold
        filtered = [img for img in images if img["similarity"] >= threshold]
        
        # Sort by similarity (highest first)
        filtered.sort(key=lambda x: x["similarity"], reverse=True)
        
        # Limit results
        return filtered[:limit]
    
    def get_statistics(self) -> Dict:
        """
        Get database statistics.
        
        Returns:
            Dict with total_images, unique_objects, unique_categories
        """
        images = self.get_all_images()
        
        if not images:
            return {
                "total_images": 0,
                "unique_objects": 0,
                "unique_categories": 0,
                "objects": [],
                "categories": []
            }
        
        objects = set(img.get("object_name", "").lower() for img in images if img.get("object_name"))
        categories = set(img.get("category", "").lower() for img in images if img.get("category"))
        
        return {
            "total_images": len(images),
            "unique_objects": len(objects),
            "unique_categories": len(categories),
            "objects": sorted(list(objects)),
            "categories": sorted(list(categories))
        }
    
    def clear_database(self):
        """Delete and recreate empty database."""
        if os.path.exists(self.path):
            os.remove(self.path)
        self.ensure_workbook()
    
    @staticmethod
    def _headers(sheet) -> List[str]:
        """Extract header row from sheet."""
        return [cell.value for cell in sheet[1] if cell.value]
